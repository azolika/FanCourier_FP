from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from LogWriter import log_writer


class FanCourierTo:
    def __init__(self):
        self.real_name = []
        self.lpl_plate = []
        self.account = []
        self.mail_address = []
        self.webeye_car_id = []
        self.wkmm1 = []
        self.can_km = []
        self.km_limit = []

    def excel_reader(self, filename, car_list_webeye, log_file_name):

        log_writer(log_file_name, "Read recipient excel start", 1)

        wb = load_workbook(filename)
        sheet_list = wb.sheetnames
        sheet = wb[sheet_list[0]]

        number_of_recipient = 1
        while 1:
            if sheet[get_column_letter(3) + str(number_of_recipient)].value is None:
                break
            else:
                number_of_recipient = number_of_recipient + 1

        i = 2
        while i < number_of_recipient:
            self.real_name.append(sheet[get_column_letter(2) + str(i)].value)
            lpl_plate = sheet[get_column_letter(3) + str(i)].value
            self.lpl_plate.append(lpl_plate)
            self.account.append(sheet[get_column_letter(4) + str(i)].value)
            email_address = sheet[get_column_letter(5) + str(i)].value

            if email_address[-1] == ";":
                email_address = email_address[:-1]
            email_address = email_address.replace(' ', '')

            self.mail_address.append(email_address.split(";"))
            self.km_limit.append(sheet[get_column_letter(6) + str(i)].value)
            try:
                car_list_webeye.car[car_list_webeye.lpl_clear.index(lpl_plate)]
            except ValueError:
                self.webeye_car_id.append(lpl_plate)
                self.wkmm1.append("")
            else:
                self.webeye_car_id.append(car_list_webeye.car[car_list_webeye.lpl_clear.index(lpl_plate)])
                self.wkmm1.append(car_list_webeye.wkmm1[car_list_webeye.lpl_clear.index(lpl_plate)])
            i = i + 1
        log_writer(log_file_name, "Read recipient excel end", 1)
